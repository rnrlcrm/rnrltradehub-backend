"""
Report generation service for exporting data to Excel, PDF, and CSV formats.

Provides functionality to export:
- Business partners
- Sales contracts
- Invoices
- Payments
- Disputes
- Commissions
- Any entity data
"""
from typing import List, Dict, Any, Optional
from datetime import datetime
import io
from sqlalchemy.orm import Session


class ReportService:
    """Service for generating reports in various formats."""
    
    def __init__(self, db: Session):
        """
        Initialize ReportService.
        
        Args:
            db: Database session
        """
        self.db = db
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> str:
        """
        Export data to CSV format.
        
        Args:
            data: List of dictionaries containing data
            columns: Optional list of column names to include
            
        Returns:
            CSV string
        """
        if not data:
            return ""
        
        # If columns not specified, use all keys from first row
        if not columns:
            columns = list(data[0].keys())
        
        # Build CSV
        csv_lines = []
        
        # Header
        csv_lines.append(','.join([f'"{col}"' for col in columns]))
        
        # Data rows
        for row in data:
            values = []
            for col in columns:
                value = row.get(col, '')
                # Handle None, datetime, and other types
                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, (list, dict)):
                    value = str(value)
                # Escape quotes and wrap in quotes
                value_str = str(value).replace('"', '""')
                values.append(f'"{value_str}"')
            csv_lines.append(','.join(values))
        
        return '\n'.join(csv_lines)
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Export data to Excel format (.xlsx).
        
        Requires: openpyxl library
        
        Args:
            data: List of dictionaries containing data
            sheet_name: Name for the Excel sheet
            columns: Optional list of column names to include
            
        Returns:
            Excel file bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxl library is required for Excel export. "
                "Install with: pip install openpyxl"
            )
        
        if not data:
            # Return empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        
        # If columns not specified, use all keys from first row
        if not columns:
            columns = list(data[0].keys())
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, '')
                
                # Handle different data types
                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, (list, dict)):
                    value = str(value)
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str = "Report",
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Export data to PDF format.
        
        Requires: reportlab library
        
        Args:
            data: List of dictionaries containing data
            title: Report title
            columns: Optional list of column names to include
            
        Returns:
            PDF file bytes
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
        except ImportError:
            raise ImportError(
                "reportlab library is required for PDF export. "
                "Install with: pip install reportlab"
            )
        
        buffer = io.BytesIO()
        
        # Create PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Add metadata
        meta_style = styles['Normal']
        meta_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Records: {len(data)}"
        elements.append(Paragraph(meta_text, meta_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        if not data:
            elements.append(Paragraph("No data available", styles['Normal']))
        else:
            # If columns not specified, use all keys from first row
            if not columns:
                columns = list(data[0].keys())
            
            # Prepare table data
            table_data = [columns]  # Header
            
            for row in data:
                row_values = []
                for col in columns:
                    value = row.get(col, '')
                    if value is None:
                        value = ''
                    elif isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, (list, dict)):
                        value = str(value)[:50] + '...' if len(str(value)) > 50 else str(value)
                    row_values.append(str(value)[:100])  # Truncate long values
                table_data.append(row_values)
            
            # Create table
            col_widths = [A4[0] / len(columns) - 0.5 * inch for _ in columns]
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style table
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def get_business_partners_data(
        self,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """
        Get business partners data for export.
        
        Args:
            filters: Optional filters (status, business_type, etc.)
            
        Returns:
            List of business partner dictionaries
        """
        from models import BusinessPartner
        
        query = self.db.query(BusinessPartner)
        
        if filters:
            if 'status' in filters:
                query = query.filter(BusinessPartner.status == filters['status'])
            if 'business_type' in filters:
                query = query.filter(BusinessPartner.business_type == filters['business_type'])
        
        partners = query.all()
        
        return [
            {
                'BP Code': bp.bp_code,
                'Legal Name': bp.legal_name,
                'Organization': bp.organization,
                'Business Type': bp.business_type,
                'Status': bp.status,
                'Contact Person': bp.contact_person,
                'Email': bp.contact_email,
                'Phone': bp.contact_phone,
                'City': bp.city,
                'State': bp.state,
                'PAN': bp.pan,
                'GSTIN': bp.gstin or '',
                'Created At': bp.created_at,
            }
            for bp in partners
        ]
    
    def get_sales_contracts_data(
        self,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """
        Get sales contracts data for export.
        
        Args:
            filters: Optional filters (status, buyer_id, seller_id, etc.)
            
        Returns:
            List of sales contract dictionaries
        """
        from models import SalesContract, BusinessPartner
        
        query = self.db.query(SalesContract).join(
            BusinessPartner,
            SalesContract.buyer_id == BusinessPartner.id
        )
        
        if filters:
            if 'status' in filters:
                query = query.filter(SalesContract.status == filters['status'])
        
        contracts = query.all()
        
        return [
            {
                'Contract Number': contract.contract_number,
                'Buyer': contract.buyer.legal_name if contract.buyer else '',
                'Seller': contract.seller.legal_name if contract.seller else '',
                'Commodity': contract.commodity,
                'Quantity': contract.quantity,
                'Unit': contract.unit,
                'Price Per Unit': contract.price_per_unit,
                'Total Value': contract.total_value,
                'Status': contract.status,
                'Start Date': contract.start_date,
                'End Date': contract.end_date,
                'Created At': contract.created_at,
            }
            for contract in contracts
        ]
    
    def get_invoices_data(
        self,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """
        Get invoices data for export.
        
        Args:
            filters: Optional filters (status, contract_id, etc.)
            
        Returns:
            List of invoice dictionaries
        """
        from models import Invoice, SalesContract
        
        query = self.db.query(Invoice)
        
        if filters:
            if 'status' in filters:
                query = query.filter(Invoice.status == filters['status'])
        
        invoices = query.all()
        
        return [
            {
                'Invoice Number': invoice.invoice_number,
                'Contract': invoice.contract.contract_number if invoice.contract else '',
                'Invoice Date': invoice.invoice_date,
                'Due Date': invoice.due_date,
                'Base Amount': invoice.base_amount,
                'GST Amount': invoice.gst_amount,
                'Total Amount': invoice.total_amount,
                'Paid Amount': invoice.paid_amount,
                'Outstanding': invoice.total_amount - (invoice.paid_amount or 0),
                'Status': invoice.status,
                'Created At': invoice.created_at,
            }
            for invoice in invoices
        ]
